import streamlit as st
import pandas as pd
import io
import base64
from file_handler import FileHandler
from openpyxl import load_workbook
from config import UPLOAD_CONFIG


class Result:
    """
    Handles the display of analysis results with tabbed interface for original data sheets
    and graphical elements.
    """
    def __init__(self):
        """Initialize with session state data"""
        self.new_df = st.session_state.get('input_excel_new', pd.DataFrame())
        self.res_df = st.session_state.get('results', pd.DataFrame())
        self.uploaded_file = st.session_state.get('new_file_object')
        
        # Configure display settings
        self.image_max_width = 800  # Maximum width for displayed images in pixels

    # ---- DATA PROCESSING METHODS ----
    
    @staticmethod
    def _highlight_row(row):
        """Apply conditional formatting to rows based on change type"""
        if row['Change Type'] == 'New':
            return ['background-color: #FF5733'] * len(row)
        if row['Change Type'] == 'Spring Changed':
            return ['background-color: #B4C6E7'] * len(row)
        return [''] * len(row)
    
    def _prepare_display_data(self):
        """Prepare the data for display by merging with results"""
        # Merge new input with metadata and sort
        display_df = self.new_df.copy()
        
        metadata_cols = [
            'Old Reference', 'New Reference',
            'Mass Status', 'Change Type',
            'Cell ID New', 'Cell ID Old'
        ]
        
        for col in metadata_cols:
            display_df[col] = self.res_df[col]
            
        return display_df.sort_values('Cell ID New', ascending=True).reset_index(drop=True)
    
    # ---- EXCEL DATA EXTRACTION METHODS ----
    
    def _get_sheets_from_excel(self):
        """
        Extract all sheets and graphical elements from the uploaded Excel file
        """
        if not self.uploaded_file:
            return {}, {}
        
        # Load the workbook from the uploaded file
        excel_data = io.BytesIO(self.uploaded_file.getvalue())
        
        # Get all sheet names using pandas (handles special characters better)
        try:
            excel_file = pd.ExcelFile(excel_data)
            sheet_names = excel_file.sheet_names
        except Exception as e:
            st.error(f"Error reading sheet names: {str(e)}")
            return {}, {}
        
        # Extract data and graphs
        sheets_data = {}
        graphs_data = {}
        
        # Reset file pointer and load workbook for image/chart extraction
        excel_data.seek(0)
        try:
            # Use data_only=True to get calculated values in cells
            wb = load_workbook(excel_data, data_only=True)
            
            # First process PTA sheet separately
            if UPLOAD_CONFIG["sheet_name"] in wb.sheetnames:
                pta_ws = wb[UPLOAD_CONFIG["sheet_name"]]
                pta_graphs = self._extract_charts_from_sheet(pta_ws)
                if pta_graphs:
                    graphs_data["PTA Graphs"] = pta_graphs
        except Exception as e:
            st.warning(f"Could not extract graphs from PTA sheet: {str(e)}")
        
        # Process all other sheets
        for sheet_name in sheet_names:
            try:
                # Skip PTA sheet - we process it separately
                if sheet_name == UPLOAD_CONFIG["sheet_name"]:
                    continue
                
                # Reset file pointer before reading each sheet
                excel_data.seek(0)
                
                # Load sheet data
                df = pd.read_excel(excel_data, sheet_name=sheet_name)
                sheets_data[sheet_name] = df
                
                # Try to extract graphs from this sheet
                if sheet_name in wb.sheetnames:
                    sheet_graphs = self._extract_charts_from_sheet(wb[sheet_name])
                    if sheet_graphs:
                        # Store graphs with descriptive name
                        graph_key = f"{sheet_name} Graphs"
                        graphs_data[graph_key] = sheet_graphs
                
            except Exception as e:
                sheets_data[sheet_name] = f"Error loading sheet: {str(e)}"
        
        return sheets_data, graphs_data
    
    def _extract_charts_from_sheet(self, worksheet):
        """Extract and process images/charts from worksheet"""
        charts = []
        
        # Process all embedded images
        try:
            for image in worksheet._images:
                try:
                    # Get image data
                    img_data = image._data()
                    if img_data:
                        # Convert to base64 for displaying in HTML
                        img_b64 = base64.b64encode(img_data).decode()
                        # Store image data with metadata for sizing
                        charts.append({
                            'data': img_b64,
                            'width': image.width,
                            'height': image.height
                        })
                except Exception as e:
                    st.warning(f"Could not extract image: {str(e)}")
        except Exception as e:
            st.warning(f"Error accessing images in worksheet: {str(e)}")
                
        return charts
    
    # ---- DISPLAY METHODS ----
    
    def display_results(self):
        """Main method to display analysis results in a tabbed interface"""
        if self.new_df.empty or self.res_df.empty:
            st.warning('No data to display.')
            return
        
        # STEP 1: Extract data and graphs from Excel file
        sheets_data, graphs_data = self._get_sheets_from_excel()
        
        # STEP 2: Create tab names in appropriate order
        tab_names = self._create_tab_names(sheets_data, graphs_data)
        
        # STEP 3: Create tabbed interface
        tabs = st.tabs(tab_names)
        
        # STEP 4: Render content in each tab
        self._render_tabs_content(tabs, tab_names, sheets_data, graphs_data)
        
        # STEP 5: Add download section below tabs
        self._add_download_section()
    
    def _create_tab_names(self, sheets_data, graphs_data):
        """Create ordered list of tab names"""
        # Always start with Analysis Results
        tab_names = ["Analysis Results"]
        
        # Add PTA Graphs next if available
        if "PTA Graphs" in graphs_data:
            tab_names.append("PTA Graphs")
            
        # Add all other sheets
        sheet_names = list(sheets_data.keys())
        
        # Try to find "Assiette théorique" or similar and place it last
        for special_name in ["Assiette théorique", "Assiette theorique", "Assiette"]:
            matching_sheets = [name for name in sheet_names if special_name.lower() in name.lower()]
            for match in matching_sheets:
                if match in sheet_names:
                    sheet_names.remove(match)
                    # Add to end
                    sheet_names.append(match)
        
        # Add all sheet names to tab_names
        tab_names.extend(sheet_names)
        
        return tab_names
    
    def _render_tabs_content(self, tabs, tab_names, sheets_data, graphs_data):
        """Render content in each tab"""
        # STEP 1: First tab - Analysis Results with highlighting
        with tabs[0]:
            self._render_analysis_results()
            
        # STEP 2: PTA Graphs tab if it exists
        tab_index = 1
        if "PTA Graphs" in graphs_data and len(tab_names) > 1:
            with tabs[tab_index]:
                self._display_graphs(graphs_data["PTA Graphs"], "PTA Sheet Graphs")
            tab_index += 1
        
        # STEP 3: Render all other sheets
        for i, sheet_name in enumerate(tab_names[tab_index:]):
            with tabs[tab_index + i]:
                # Display sheet data if available
                if sheet_name in sheets_data:
                    sheet_content = sheets_data[sheet_name]
                    if isinstance(sheet_content, pd.DataFrame):
                        st.dataframe(sheet_content)
                    else:
                        st.error(sheet_content)  # Display error message
                    
                    # Special handling for last sheet which might be Assiette théorique
                    is_last_sheet = (tab_index + i == len(tab_names) - 1)
                    special_sheet = any(s in sheet_name.lower() for s in ["assiette", "théorique", "theorique"])
                    
                    # Add graphs from this sheet if available
                    graph_key = f"{sheet_name} Graphs"
                    if graph_key in graphs_data:
                        title = "Sheet Graphs"
                        if is_last_sheet and special_sheet:
                            title = "Assiette Théorique"
                            # Special display for this important graph
                            self._display_graphs(graphs_data[graph_key], title, is_special=True)
                        else:
                            self._display_graphs(graphs_data[graph_key], title)
                
                # Add sheet name caption
                st.caption(f"Sheet: {sheet_name}")
    
    def _render_analysis_results(self):
        """Render the analysis results with styling in the first tab"""
        # Prepare data
        display_df = self._prepare_display_data()
        
        # Apply styling
        styled = display_df.style.apply(self._highlight_row, axis=1)
        
        # Display styled dataframe
        st.dataframe(styled, use_container_width=True)
        
        # Add color legend
        st.markdown('''
        **Color Legend:**
        - 🟥 **New rows**: Cars added in new PTA file
        - 🟦 **Spring Changed**: Reference (spring) changed
        ''')
    
    def _display_graphs(self, graphs, title=None, is_special=False):
        """Display graphs with appropriate sizing and formatting"""
        if not graphs:
            return
            
        if title:
            st.subheader(title)
            
        for i, img_info in enumerate(graphs):
            try:
                img_b64 = img_info['data'] if isinstance(img_info, dict) else img_info
                
                # Calculate display size
                if isinstance(img_info, dict) and 'width' in img_info and 'height' in img_info:
                    # Get original dimensions
                    orig_width = img_info['width']
                    orig_height = img_info['height']
                    
                    # Calculate appropriate display width
                    display_width = min(self.image_max_width, orig_width)
                    
                    # Scale height proportionally
                    if orig_width > 0:  # Avoid division by zero
                        scale_factor = display_width / orig_width
                        display_height = int(orig_height * scale_factor)
                    else:
                        display_height = orig_height
                    
                    # Special handling for important graphs
                    if is_special:
                        # Make important graphs more prominent
                        display_width = min(self.image_max_width, int(orig_width * 1.2))
                        if orig_width > 0:
                            scale_factor = display_width / orig_width
                            display_height = int(orig_height * scale_factor)
                else:
                    # Default size if dimensions not available
                    display_width = "100%"
                    display_height = None
                
                # Display image with appropriate size
                if display_height:
                    st.markdown(f"""
                    <div style="text-align: center; margin: 20px auto;">
                        <img src="data:image/png;base64,{img_b64}" 
                             style="width: {display_width}px; height: {display_height}px; max-width: 100%;">
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="text-align: center; margin: 20px auto;">
                        <img src="data:image/png;base64,{img_b64}" 
                             style="width: {display_width}; max-width: 100%;">
                    </div>
                    """, unsafe_allow_html=True)
            except Exception as e:
                st.warning(f"Could not display image {i+1}: {str(e)}")
    
    def _add_download_section(self):
        """Add download section for Excel report"""
        st.subheader('📥 Download Results')
        try:
            # Prepare data for export
            display_df = self._prepare_display_data()
            
            # Generate Excel bytes
            data = FileHandler.create_excel_bytes(display_df)
            
            # Add download button
            st.download_button(
                '📄 Download Excel Report',
                data=data,
                file_name='spring_change_analysis.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            st.error(f'Error creating Excel file: {e}')
