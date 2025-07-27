"""
This script will validate the excel file uploaded by the user then validate the crucial columns
after that creating the excel output when the user press on download button
"""
import streamlit as st
#__TODO: import libraries_______________________________________________
import io
from typing import Any, Tuple, Optional
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from config import UPLOAD_CONFIG, REQUIRED_COLUMNS

class FileHandler:
    """Handles validation and export of Excel files."""

    #__TODO: Validate the uploaded excel buffer_______________________________________________
    @staticmethod
    def validate_excel_file(
        file: Any, file_label: str
    ) -> Tuple[bool, str, Optional[pd.DataFrame]]:
        """
        Args:
            file: Uploaded file.
            file_label: A label for the file (e.g., "old", "new").

        Returns:
            Tuple containing:
              - validity (bool)
              - message (str)
              - DataFrame if valid, else None
        """
        if not file:
            return False, f"No '{file_label}' file uploaded.", None

        try:
            df = (
                pd.read_excel(
                    file,
                    engine="openpyxl",
                    sheet_name=UPLOAD_CONFIG["sheet_name"],
                    skiprows=UPLOAD_CONFIG["skip_rows"],
                )
                .reset_index(drop=True)
            )
        except Exception as e:
            return False, f"Error reading '{file_label}' file: {e}", None

        if df.empty:
            return False, f"'{file_label}' file is empty.", None

        is_valid, msg = FileHandler._validate_columns(df)
        if not is_valid:
            return False, msg, None

        return True, "File uploaded successfully.", df
    
    #__TODO: Validate the crucial columns_______________________________________________
    @staticmethod
    def _validate_columns(df: pd.DataFrame) -> Tuple[bool, str]:
        """
        Check for required columns

        Args:
            df: DataFrame to validate.

        Returns:
            validity and error message if invalid.
        """
        required_cols = [
            REQUIRED_COLUMNS["mass"],
            REQUIRED_COLUMNS["reference"],
        ]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            return False, f"Missing columns: {', '.join(missing)}."
        return True, ""
    
    #__TODO: Create the excel output _______________________________________________
    @staticmethod
    def create_excel_bytes(
        results: pd.DataFrame
    ) -> bytes:
        """
        Generate a complete Excel report as bytes that preserves all sheets from the 
        original PTA file while highlighting changes in the PTA sheet.

        Args:
            results: Analysis results with metadata.

        Returns:
            Byte content of the Excel file.
        """
        # Get the original uploaded file - use the correct session state key
        uploaded_file = st.session_state.get('new_file_object')
        results_df = st.session_state.get('results')
        
        if uploaded_file is None or results_df is None:
            raise ValueError("Both 'results' and 'original file' are required.")
        
        # Create a BytesIO object to hold the workbook
        output = io.BytesIO()
        
        # First, save the uploaded file to a temporary BytesIO object
        temp_io = io.BytesIO()
        temp_io.write(uploaded_file.getvalue())
        temp_io.seek(0)
        
        # Load the workbook from the BytesIO object
        wb = load_workbook(temp_io)
        
        # Get the PTA sheet
        if UPLOAD_CONFIG["sheet_name"] in wb.sheetnames:
            ws = wb[UPLOAD_CONFIG["sheet_name"]]
            
            # Find data start row (after skipping header rows)
            start_row = UPLOAD_CONFIG["skip_rows"][0] + 2  # Skip header + extra row
            
            # Try to find the reference column based on header row content
            reference_col = None
            headers_found = []
            
            # Scan header row to find reference column
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=start_row-1, column=col_idx).value
                if cell_value:
                    headers_found.append(str(cell_value).strip())
                    if REQUIRED_COLUMNS["reference"] in str(cell_value):
                        reference_col = col_idx
                        break
            
            # If not found, try with more flexible matching
            if reference_col is None:
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=start_row-1, column=col_idx).value
                    if cell_value and "référence" in str(cell_value).lower():
                        reference_col = col_idx
                        break
            
            # If still not found, proceed without reference column validation
            # Just use Cell IDs from results dataframe for matching rows
            
            # Match each row with the results and apply highlighting
            row_idx = start_row
            while ws.cell(row=row_idx, column=1).value is not None:
                # Get cell ID (Excel row number)
                cell_id = row_idx
                
                # Find matching record in results dataframe by cell ID
                match = results_df[results_df['Cell ID New'] == cell_id]
                
                if not match.empty:
                    change_type = match['Change Type'].values[0]
                    if change_type == 'New':
                        # Highlight new rows
                        fill = PatternFill('solid', fgColor='FF5733')
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col).fill = fill
                    elif change_type == 'Spring Changed':
                        # Highlight spring changed rows
                        fill = PatternFill('solid', fgColor='B4C6E7')
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col).fill = fill
                
                row_idx += 1
        
        # Save the workbook to the BytesIO object
        wb.save(output)
        output.seek(0)
        return output.getvalue()